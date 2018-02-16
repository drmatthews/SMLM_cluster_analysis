#!/usr/bin/env python2
# -*- coding: utf-8 -*-
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from read_roi import read_roi_file
from read_roi import read_roi_zip

from localisations import Localisations
from cluster.voronoi_ import voronoi, voronoi_segmentation
from cluster.utils import (optics_clustering,
                           kdtree_nn,
                           plot_optics_clusters,
                           polygon_area,
                           polygon_perimeter)
from cluster import statistics

import warnings
warnings.filterwarnings("ignore")

def collate_dataset_stats(out_dir, dataset_name, dataset_type, sheetname):
    print(dataset_name)
    df_list = []
    for filename in os.listdir(out_dir):
        if filename.endswith('xlsx'):
            if (dataset_type in filename) and (dataset_name in filename):
                print(filename)
                df_list.append(pd.read_excel(
                    os.path.join(out_dir,filename),
                    sheetname=sheetname))

    return pd.concat(df_list)

def cluster_near_neighbours(out_dir, dataset_name, dataset_type):
    print(dataset_name)
    nn_list = []
    for filename in os.listdir(out_dir):
        if filename.endswith('xlsx'):
            if (dataset_type in filename) and (dataset_name in filename):
                print(filename)
                df = pd.read_excel(
                                os.path.join(out_dir,filename),
                                sheetname='cluster coordinates')
                nn = kdtree_nn(df.as_matrix(columns=['x [nm]', 'y [nm]']))
                nn_list.append(nn)

    nn_df = pd.DataFrame(np.concatenate(nn_list, axis=0))
    nn_df.columns = ['Distance [nm]']

    return nn_df

def run_optics(parameters, verbose=True, use_roi=False, pixel_size=16.0):

    # parameters
    eps = parameters['eps']
    eps_extract = parameters['eps_extract']
    min_samples = parameters['min_samples']
    input_dir = parameters['input_dir']
    output_dir = parameters['output_dir']
    data_source = parameters['data_source']

    for filename in os.listdir(input_dir):

        if 'nstorm' in data_source:
            ext = 'txt'
        elif 'thunderstorm' in data_source:
            ext = 'csv'

        file_ext = os.path.splitext(filename)[1]
        if ext in file_ext:
            print("optics clustering for file %s"%filename)
            filepath = os.path.join(input_dir,filename)
            basename = os.path.splitext(filename)[0]
            output_path = os.path.join(output_dir, basename+'_optics.xlsx')

            data = Localisations(filepath, source=data_source)
            points = data.points
            xy = points.as_matrix(columns=['x','y'])

            if use_roi:
                roi_zip_path = os.path.join(input_dir, basename + '_roiset.zip')
                roi_file_path = os.path.join(input_dir, basename + '_roiset.roi')
                if os.path.exists(roi_zip_path):
                    rois = read_roi_zip(roi_zip_path)
                elif os.path.exists(roi_file_path):
                    rois = read_roi_file(roi_file_path)
                else:
                    raise ValueError(("No ImageJ roi file exists -"
                                     "you should put the file in the same directory as the data"))

                coords = {}
                for roi_id, roi in rois.items():
                    for k, v in roi.items():
                        if not isinstance(v, str):
                            roi[k] = float(v) * pixel_size

                    coords[roi_id] = xy[(xy[:, 0] > roi['left']) &
                                        (xy[:, 0] < roi['left'] + roi['width']) &
                                        (xy[:, 1] > roi['top']) &
                                        (xy[:, 1] < roi['top'] + roi['height'])]
            else:
                coords = dict([('image', xy)])

            (optics_clusters, noise) = optics_clustering(coords,
                                                         eps=eps,
                                                         eps_extract=eps_extract,
                                                         min_samples=min_samples)

            if verbose:
                if optics_clusters:
                    for roi_id in optics_clusters.keys():
                        plot_filename = os.path.join(output_dir, basename + '_{0}_clusters.png'.format(roi_id))
                        if noise:
                            noise_in_roi = noise[roi_id]
                        else:
                            noise_in_roi = None
                        plot_optics_clusters(optics_clusters[roi_id],
                                             noise=noise_in_roi,
                                             save=True,
                                             filename=plot_filename)

            for roi in optics_clusters.keys():

                oc = optics_clusters[roi]
                if oc.n_clusters > 0:
                    oc.save(output_path, sheetname=roi)
                else:
                    print("No clusters found in %s"%filename)
                    continue


def run_voronoi(parameters, verbose=True, use_roi=False, pixel_size=16.0):

    # parameters
    density_factor = parameters['density_factor']
    min_samples = parameters['min_samples']
    input_dir = parameters['input_dir']
    output_dir = parameters['output_dir']

    for filename in os.listdir(input_dir):

        if filename.endswith('csv'):
            print("voronoi clustering for file %s"%filename)
            filepath = os.path.join(input_dir,filename)
            basename = os.path.splitext(filename)[0]
            output_path = os.path.join(output_dir, basename+'_voronoi.xlsx')

            locs = pd.read_csv(filepath)

            if use_roi:
                roi_zip_path = os.path.join(input_dir, basename + '_roiset.zip')
                roi_file_path = os.path.join(input_dir, basename + '_roiset.roi')
                if os.path.exists(roi_zip_path):
                    rois = read_roi_zip(roi_zip_path)
                elif os.path.exists(roi_file_path):
                    rois = read_roi_file(roi_file_path)
                else:
                    raise ValueError(("No ImageJ roi file exists -"
                                     "you should put the file in the same directory as the data"))

                coords = {}
                for roi_id, roi in rois.items():
                    for k, v in roi.items():
                        if not isinstance(v, str):
                            roi[k] = float(v) * pixel_size

                    coords[roi_id] = locs[(locs['x [nm]'] > roi['left']) &
                                          (locs['x [nm]'] < roi['left'] + roi['width']) &
                                          (locs['y [nm]'] > roi['top']) &
                                          (locs['y [nm]'] < roi['top'] + roi['height'])].reset_index(drop=True)
            else:
                coords = dict([('image', locs)])

            for roi_id in coords.keys():
                df = coords[roi_id]

                if len(df.index) == 0:
                    print("no coords in roi")
                    continue

                voronoi_df = voronoi(df, density_factor, min_samples, show_plot=verbose, verbose=verbose)

                cluster_locs_df = voronoi_df[voronoi_df['lk'] != -1]
                labels = cluster_locs_df['lk'].unique()
                cluster_stats = {}
                cluster_stats['area'] = []
                cluster_stats['occupancy'] = []
                for m in labels:
                    cluster = cluster_locs_df[cluster_locs_df.lk == m]
                    cluster_stats['area'].append(cluster['area'].sum())
                    cluster_stats['occupancy'].append(len(cluster.index))

                cluster_stats_df = pd.DataFrame(cluster_stats)
                if any(labels):
                    writer = pd.ExcelWriter(output_path, engine='openpyxl')
                    if os.path.exists(output_path):
                        book = load_workbook(output_path)
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                    cluster_locs_df.to_excel(
                        writer,
                        sheet_name='{0} localisations'.format(roi_id),
                        index=False
                    )
                    cluster_stats_df.to_excel(
                        writer,
                        sheet_name='{0} voronoi stats'.format(roi_id),
                        index=False
                    )
                    writer.save()
                else:
                    print("No clusters found in %s"%filename)
                    continue


def collect_stats(df, col):
    total = float(df.shape[0])
    not_noise = df[df[col] != -1]
    percentage = float(not_noise.shape[0]) / total
    objects_with_clusters = (
        float(df[(df['object_id'] != -1) & (df['cluster_id'] != -1)].shape[0])
    )
    labels = not_noise[col].unique()
    area = []
    perimeter = []
    occupancy = []
    stats = {}
    stats['labels'] = labels
    for m in labels:
        group = df[df[col] == m]
        points = group.as_matrix(['x [nm]', 'y [nm]'])
        area.append(polygon_area(points))
        perimeter.append(polygon_perimeter(points))
        occupancy.append(len(group.index))

    stats['area'] = area
    stats['perimeter'] = perimeter
    stats['occupancy'] = occupancy
    stats['percentage {0}s'.format(col[:len(col) - 3])] = percentage
    stats['not_noise'] = float(not_noise.shape[0])
    stats['objects_with_clusters'] = objects_with_clusters
    return stats


def run_voronoi_segmentation(parameters,
                             use_roi=False,
                             segment_rois=False,
                             monte_carlo=True):

    # parameters
    odf = parameters['object_density_factor']
    oms = parameters['object_min_samples']
    cdf = parameters['cluster_density_factor']
    cms = parameters['cluster_min_samples']
    conditions = parameters['conditions']
    input_dir = parameters['input_dir']
    output_dir = parameters['output_dir']

    stats = {}
    for filename in os.listdir(input_dir):

        if filename.endswith('csv'):
            print("voronoi clustering for file {}".format(filename))
            locs_path = os.path.join(input_dir, filename)
            basename = os.path.splitext(filename)[0]
            output_path = os.path.join(output_dir, basename+'_voronoi.xlsx')

            roi_path = None
            if use_roi:
                roi_zip_path = os.path.join(input_dir, basename + '_roiset.zip')
                roi_file_path = os.path.join(input_dir, basename + '_roiset.roi')
                if os.path.exists(roi_zip_path):
                    roi_path = roi_zip_path
                elif os.path.exists(roi_file_path):
                    roi_path = roi_file_path
                else:
                    raise ValueError(("No ImageJ roi file exists -"
                                     "you should put the file in the same directory"
                                     "as the data"))

            vrois = voronoi_segmentation(locs_path,
                                         roi_path=roi_path,
                                         segment_rois=segment_rois,
                                         monte_carlo=monte_carlo,
                                         object_density_factor=odf,
                                         object_min_samples=oms,
                                         cluster_density_factor=cdf,
                                         cluster_min_samples=cms,
                                         show_plot=False)

            roi_cluster_means = []
            for vr_id, vroi in vrois.items():
                v_locs = vroi['locs']
                object_stats = collect_stats(v_locs, 'object_id')
                cluster_stats = collect_stats(v_locs, 'cluster_id')
                cluster_stats['percentage objects'] = object_stats['percentage objects']
                poc = (
                    object_stats['objects_with_clusters'] / object_stats['not_noise']
                )
                cluster_stats['percentage_objects_with_clusters'] = poc

                object_stats_df = pd.DataFrame(object_stats)
                cluster_stats_df = pd.DataFrame(cluster_stats)
                print(cluster_stats_df.head())
                # remove unwanted columns
                cluster_stats_df.drop(columns=['objects_with_clusters', 'not_noise'])
                # reorder ready for saving
                cluster_stats_df = cluster_stats_df[['area',
                                                     'perimeter',
                                                     'occupancy',
                                                     'percentage objects',
                                                     'percentage clusters',
                                                     'percentage_objects_with_clusters']]
                # collect the mean values for each parameter in df to write
                # to file later - list of pandas series
                roi_cluster_means.append(cluster_stats_df.mean())

                writer = pd.ExcelWriter(output_path, engine='openpyxl')
                if os.path.exists(output_path):
                    book = load_workbook(output_path)
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                v_locs.to_excel(
                    writer,
                    sheet_name='{0} localisations'.format(vr_id),
                    index=False
                )
                object_stats_df.to_excel(
                    writer,
                    sheet_name='{0} object stats'.format(vr_id),
                    index=False
                )
                cluster_stats_df.to_excel(
                    writer,
                    sheet_name='{0} cluster stats'.format(vr_id),
                    index=False
                )
                writer.save()

            # turn list of series into a dataframe
            if len(roi_cluster_means) > 1:
                stats[filename] = pd.concat(roi_cluster_means,
                                            axis=0,
                                            keys=[s.name for s in roi_cluster_means])
            else:
                stats[filename] = pd.DataFrame(roi_cluster_means[0]).T

    print("stats: {0}".format(stats))
    fnames = list(stats.keys())
    stats_path = os.path.join(output_dir, 'cluster_statistics.xlsx')
    for condition in conditions:
        condition_fnames = [fname for fname in fnames if condition in fname]
        data = {'files': condition_fnames}
        for cf in condition_fnames:
            cluster_stats = stats[cf]
            for column in cluster_stats:
                if column != 'labels':
                    data[column] = cluster_stats[column].mean()

        statistics.write_stats(stats_path, data, condition)

if __name__=='__main__':
    parameters = {}
    parameters['pixel_size'] = 16.0 #nm
    parameters['object_min_samples'] = 3
    parameters['object_density_factor'] = 4.213
    parameters['cluster_min_samples'] = 3
    parameters['cluster_density_factor'] = 20
    parameters['conditions'] = ['wtTNF', 'control']
    parameters['input_dir'] = "C:\\Users\\NIC ADMIN\\Documents\\atto647n pre-hawk"
    parameters['output_dir'] = "C:\\Users\\NIC ADMIN\\Documents\\atto647n pre-hawk\\processed\\monte_carlo"
    parameters['data_source'] = 'thunderstorm'
    run_voronoi_segmentation(parameters, use_roi=False, segment_rois=False)

    # parameters = {}
    # parameters['pixel_size'] = 16.0 #nm
    # parameters['object_min_samples'] = 3
    # parameters['cluster_min_samples'] = 3
    # parameters['cluster_density_factor'] = 20
    # parameters['input_dir'] = "C:\\Users\\NIC ADMIN\\Documents\\feb18 dSTORM"
    # parameters['output_dir'] = "C:\\Users\\NIC ADMIN\\Documents\\feb18 dSTORM\\processed"
    # parameters['data_source'] = 'thunderstorm'
    # run_voronoi_segmentation(parameters, use_roi=False, segment_rois=False)
